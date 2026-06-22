#!/usr/bin/env python3
"""
Ferncrest Data Pipeline
=======================
Reads Cloudbeds occupancy Excel exports and writes actuals + pace data
into the property workbook (ferncrest_ok-fc_v2.xlsx).

Usage:
    python ferncrest_pipeline.py

File expectations (set paths in CONFIG below):
    - ok-fc_occupancy_2026ytd.xlsx   : Jan 1 – Apr 30 actuals
    - ok-fc_occupancy_pace_90day.xlsx: May 1 – Jul 30 forward pace
    - ferncrest_ok-fc_v2.xlsx        : property workbook

What it does:
    1. Parses both Cloudbeds exports
    2. Aggregates daily rows → monthly totals
    3. Calculates ADR, occupancy %, RevPAR per month
    4. Writes actuals into Actuals_Input tab (Jan–Apr)
    5. Writes pace data into Pace_Weekly tab (May–Jul)
    6. Saves workbook
    7. Prints a summary to console (Slack/report step comes next)
"""

import openpyxl
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime, date, timedelta
import os
import sys
import json
from dotenv import load_dotenv

# Detect if running on GitHub Actions
_GDRIVE = "/Users/stephcleung/Library/CloudStorage/GoogleDrive-stephanie@lintonhospitality.com/Shared drives/Ferncrest/1_Locations/02_OK-FC/2_Finance AM/AssetManagement"
_CI = os.environ.get("CI") == "true"
_BASE = os.path.dirname(os.path.abspath(__file__)) if _CI else _GDRIVE

load_dotenv(os.path.join(_BASE, ".env"), override=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG — paths auto-switch between local Mac and GitHub Actions
# ─────────────────────────────────────────────────────────────────────────────
CONFIG = {
    "ytd_file":       os.path.join(_BASE, "ok-fc_occupancy_2026ytd.xlsx"),
    "pace_file":      os.path.join(_BASE, "ok-fc_occupancy_pace_150day.xlsx"),
    "workbook":       os.path.join(_BASE, "ferncrest_ok-fc_v2.xlsx"),
    "snapshot_cache": os.path.join(_BASE, "pace_snapshot_cache.json"),
    "full_report_url": "https://linton-hospitality.github.io/FlintCreekReport/",
    "property_code":  "OK-FC",
    "total_units":    12,
    "fiscal_year":    2026,
}

# ── OK-FC pace targets (Slack digest) — update seasonally ──────────────────
# "weeks_out: fraction of target occupancy that should be booked by then"
PACE_TARGETS_OKFC = {5: 0.10, 4: 0.25, 3: 0.35, 2: 0.50, 1: 0.65}

# Target occupancy by arrival month (drives the forward pace table)
TARGET_OCC_OKFC = {
    1: 0.15, 2: 0.15, 3: 0.25, 4: 0.20,
    5: 0.25, 6: 0.35, 7: 0.50, 8: 0.50,
    9: 0.30, 10: 0.40, 11: 0.25, 12: 0.20,
}

# Days in each month for 2026 (non-leap year)
DAYS_IN_MONTH = {
    1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30,
    7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31
}

MONTH_NAMES = {
    1:"Jan", 2:"Feb", 3:"Mar", 4:"Apr", 5:"May", 6:"Jun",
    7:"Jul", 8:"Aug", 9:"Sep", 10:"Oct", 11:"Nov", 12:"Dec"
}

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: Parse a Cloudbeds occupancy export
# ─────────────────────────────────────────────────────────────────────────────
def parse_cloudbeds_export(filepath):
    """
    Reads a Cloudbeds occupancy xlsx export.
    Returns a dict keyed by month number (int):
        {
            1: {
                "nights_sold": 56,
                "revenue": 13503.44,
                "available_nights": 372,   # units × days in month
                "occupancy_pct": 0.151,
                "adr": 241.13,
                "revpar": 36.30,
                "daily_rows": [...]         # raw daily data for pace tab
            },
            ...
        }
    """
    print(f"\n📂 Reading: {filepath}")

    if not os.path.exists(filepath):
        print(f"  ❌ File not found: {filepath}")
        sys.exit(1)

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Validate headers
    headers = rows[0]
    expected = ("Date", "ADR 2026",
                "ACCOMMODATIONS BOOKED 2026", "Room Rates 2026")
    if headers != expected:
        print(f"  ⚠️  Unexpected headers: {headers}")
        print(f"       Expected: {expected}")
        print(f"       Column mapping may be off — check before trusting output.")

    # Aggregate by month
    monthly = defaultdict(lambda: {
        "nights_sold": 0,
        "revenue": 0.0,
        "daily_rows": []
    })

    skipped = 0
    for row in rows[1:]:
        date_str, adr_val, booked, revenue = row

        # Skip blank or zero rows (no-bookings days still matter for
        # available nights calc, but zero revenue/bookings rows are fine)
        if date_str is None:
            skipped += 1
            continue

        # Parse MM/DD date format from Cloudbeds
        try:
            month = int(str(date_str).split("/")[0])
        except (ValueError, IndexError):
            skipped += 1
            continue

        nights = int(booked) if booked else 0
        rev    = float(revenue) if revenue else 0.0

        monthly[month]["nights_sold"] += nights
        monthly[month]["revenue"]     += rev
        monthly[month]["daily_rows"].append({
            "date":        date_str,
            "adr":         float(adr_val) if adr_val else 0.0,
            "nights_sold": nights,
            "revenue":     rev,
        })

    if skipped:
        print(f"  ℹ️  Skipped {skipped} blank rows")

    # Calculate derived metrics per month
    result = {}
    units = CONFIG["total_units"]
    for month, data in sorted(monthly.items()):
        days        = DAYS_IN_MONTH.get(month, 30)
        avail       = units * days
        nights_sold = data["nights_sold"]
        revenue     = data["revenue"]
        occ         = nights_sold / avail if avail > 0 else 0
        adr         = revenue / nights_sold if nights_sold > 0 else 0
        revpar      = revenue / avail if avail > 0 else 0

        result[month] = {
            "nights_sold":      nights_sold,
            "available_nights": avail,
            "revenue":          round(revenue, 2),
            "occupancy_pct":    round(occ, 4),
            "adr":              round(adr, 2),
            "revpar":           round(revpar, 2),
            "daily_rows":       data["daily_rows"],
        }

        print(f"  {MONTH_NAMES[month]:>3}: "
              f"nights={nights_sold:>3} / {avail}  "
              f"occ={occ:>5.1%}  "
              f"ADR=${adr:>7.2f}  "
              f"rev=${revenue:>9,.2f}")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: Write actuals into Actuals_Input tab
# ─────────────────────────────────────────────────────────────────────────────
def write_actuals(ws_act, monthly_data):
    """
    Finds the correct rows in Actuals_Input and writes monthly values.
    Columns D–O = Jan–Dec (columns 4–15).
    """
    print("\n📝 Writing actuals into Actuals_Input tab...")

    # Row map: label → row number in Actuals_Input
    # We scan the sheet to find them dynamically — safer than hardcoding
    row_map = {}
    for row in ws_act.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                label = cell.value.strip()
                if label in [
                    "Room Revenue (excl. tax)",
                    "Nights Available",
                    "Nights Occupied",
                    "ADR (Actual)",
                ]:
                    row_map[label] = cell.row

    if not row_map:
        print("  ⚠️  Could not find target rows in Actuals_Input — check tab structure")
        return

    print(f"  Found rows: {row_map}")

    for month, data in monthly_data.items():
        col = 3 + month  # Jan=col4, Feb=col5, etc.

        if "Room Revenue (excl. tax)" in row_map:
            ws_act.cell(row_map["Room Revenue (excl. tax)"], col,
                        data["revenue"])

        if "Nights Available" in row_map:
            ws_act.cell(row_map["Nights Available"], col,
                        data["available_nights"])

        if "Nights Occupied" in row_map:
            ws_act.cell(row_map["Nights Occupied"], col,
                        data["nights_sold"])

        if "ADR (Actual)" in row_map:
            ws_act.cell(row_map["ADR (Actual)"], col,
                        data["adr"])

        print(f"  ✅ {MONTH_NAMES[month]}: "
              f"rev=${data['revenue']:,.2f}  "
              f"nights={data['nights_sold']}  "
              f"ADR=${data['adr']:.2f}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: Write pace data into Pace_Weekly tab
# ─────────────────────────────────────────────────────────────────────────────
def write_pace(ws_pace, monthly_data, as_of_date=None):
    """
    Appends a new weekly snapshot block to the Pace_Weekly rolling log.
    Finds the last data row and appends below it.
    """
    print("\n📊 Writing pace data into Pace_Weekly tab...")

    if as_of_date is None:
        as_of_date = date.today().strftime("%Y-%m-%d")

    # Budget occupancy assumptions (from Config/Budget tab)
    budget_occ = {
        5: 0.30, 6: 0.30, 7: 0.40, 8: 0.55,
        9: 0.20, 10: 0.50, 11: 0.20, 12: 0.10
    }

    # Find last used row in Pace_Weekly
    last_row = 1
    for row in ws_pace.iter_rows():
        for cell in row:
            if cell.value is not None:
                last_row = max(last_row, cell.row)

    # Add a section header for this week's snapshot
    insert_row = last_row + 2
    ws_pace.merge_cells(
        start_row=insert_row, start_column=1,
        end_row=insert_row, end_column=14
    )
    header_cell = ws_pace.cell(insert_row, 1,
        f"▶  WEEK OF {as_of_date}  (script auto-populated)")
    header_cell.font = openpyxl.styles.Font(
        name="Arial", bold=True, color="FFFCF9", size=10)
    header_cell.fill = openpyxl.styles.PatternFill(
        "solid", start_color="FF1F3320", end_color="FF1F3320")
    header_cell.alignment = openpyxl.styles.Alignment(
        horizontal="left", vertical="center", indent=1)

    data_row = insert_row + 1
    units = CONFIG["total_units"]

    for month, data in sorted(monthly_data.items()):
        days        = DAYS_IN_MONTH.get(month, 30)
        avail       = units * days
        nights_sold = data["nights_sold"]
        bud_occ     = budget_occ.get(month, 0.25)
        month_label = f"{MONTH_NAMES[month]} {CONFIG['fiscal_year']}"

        # Weeks until arrival (approx)
        today = date.today()
        # First of the arrival month
        arr_date = date(CONFIG["fiscal_year"], month, 1)
        weeks_out = max(0, (arr_date - today).days // 7)

        ws_pace.cell(data_row, 1,  as_of_date)
        ws_pace.cell(data_row, 2,  month_label)
        ws_pace.cell(data_row, 3,  weeks_out)
        ws_pace.cell(data_row, 4,  avail)
        ws_pace.cell(data_row, 5,  nights_sold)
        # Col 6: occ % paced — formula
        ws_pace.cell(data_row, 6,
            f"=IFERROR(E{data_row}/D{data_row},0)")
        ws_pace.cell(data_row, 6).number_format = "0.0%"
        # Col 7: LY same point — leave blank for now (no LY data yet)
        ws_pace.cell(data_row, 7,  0)
        # Col 8: LY occ %
        ws_pace.cell(data_row, 8,
            f"=IFERROR(G{data_row}/D{data_row},0)")
        ws_pace.cell(data_row, 8).number_format = "0.0%"
        # Col 9: Budget occ %
        ws_pace.cell(data_row, 9,  bud_occ)
        ws_pace.cell(data_row, 9).number_format = "0.0%"
        # Col 10: Pace vs LY
        ws_pace.cell(data_row, 10,
            f"=IFERROR(F{data_row}-H{data_row},0)")
        ws_pace.cell(data_row, 10).number_format = "0.0%"
        # Col 11: Pace vs Budget
        ws_pace.cell(data_row, 11,
            f"=IFERROR(F{data_row}-I{data_row},0)")
        ws_pace.cell(data_row, 11).number_format = "0.0%"
        # Col 12: ADR on books
        ws_pace.cell(data_row, 12, data["adr"])
        ws_pace.cell(data_row, 12).number_format = '$#,##0.00'
        # Col 13: Pickup since last week
        ws_pace.cell(data_row, 13,
            f'=IFERROR(E{data_row}-VLOOKUP(B{data_row},'
            f'B$5:E{data_row-1},4,0),"— first entry")')
        # Col 14: Flag
        ws_pace.cell(data_row, 14,
            f'=IF(E{data_row}=0,"— no bookings yet",'
            f'IF(K{data_row}<-0.05,"⚠️ Behind budget — review rate",'
            f'IF(K{data_row}>0.05,"✅ Ahead — hold rate or push ADR",'
            f'IF(J{data_row}<-0.03,"↓ Soft vs LY — monitor",'
            f'"◼ On track"))))')

        occ_pct = nights_sold / avail if avail > 0 else 0
        vs_bud  = occ_pct - bud_occ
        flag    = ("⚠️ Behind" if vs_bud < -0.05
                   else "✅ Ahead" if vs_bud > 0.05
                   else "◼ On track")

        print(f"  {month_label}: {nights_sold} nights / {avail} avail  "
              f"({occ_pct:.1%} paced vs {bud_occ:.1%} budget)  {flag}")

        data_row += 1


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: Update Report_Data bridge tab
# ─────────────────────────────────────────────────────────────────────────────
def update_report_data(ws_rd, ytd_data, pace_data):
    """
    Updates key-value pairs in Report_Data tab so the HTML report
    script can read a single structured output without touching
    other tabs.
    """
    print("\n🔗 Updating Report_Data bridge tab...")

    # Build update dict: key → new value
    # Aggregate YTD totals
    ytd_revenue     = sum(d["revenue"] for d in ytd_data.values())
    ytd_nights      = sum(d["nights_sold"] for d in ytd_data.values())
    ytd_avail       = sum(d["available_nights"] for d in ytd_data.values())
    ytd_occ         = ytd_nights / ytd_avail if ytd_avail else 0
    ytd_adr         = ytd_revenue / ytd_nights if ytd_nights else 0
    ytd_revpar      = ytd_revenue / ytd_avail if ytd_avail else 0

    # Pace: May, Jun, Jul on books
    may = pace_data.get(5, {})
    jun = pace_data.get(6, {})
    jul = pace_data.get(7, {})

    updates = {
        "actual_net_revenue_ytd":   round(ytd_revenue, 2),
        "actual_nights_sold_ytd":   ytd_nights,
        "actual_occ_rate_ytd":      round(ytd_occ, 4),
        "actual_adr_ytd":           round(ytd_adr, 2),
        "actual_revpar_ytd":        round(ytd_revpar, 2),
        "pace_may_on_books":        may.get("nights_sold", 0),
        "pace_may_occ_pct":         may.get("occupancy_pct", 0),
        "pace_may_adr":             may.get("adr", 0),
        "pace_jun_on_books":        jun.get("nights_sold", 0),
        "pace_jun_occ_pct":         jun.get("occupancy_pct", 0),
        "pace_jun_adr":             jun.get("adr", 0),
        "pace_jul_on_books":        jul.get("nights_sold", 0),
        "pace_jul_occ_pct":         jul.get("occupancy_pct", 0),
        "pace_jul_adr":             jul.get("adr", 0),
        "last_updated":             date.today().strftime("%Y-%m-%d"),
    }

    # Scan Report_Data for matching keys and update values
    updated = []
    for row in ws_rd.iter_rows():
        key_cell = row[0]  # Column A
        val_cell = row[1]  # Column B
        if key_cell.value in updates:
            val_cell.value = updates[key_cell.value]
            updated.append(key_cell.value)

    # Add any missing keys at the bottom
    missing = [k for k in updates if k not in updated]
    if missing:
        last = ws_rd.max_row + 1
        for key in missing:
            ws_rd.cell(last, 1, key)
            ws_rd.cell(last, 2, updates[key])
            last += 1
        print(f"  ℹ️  Added {len(missing)} new keys: {missing}")

    print(f"  ✅ Updated {len(updated)} Report_Data fields")
    print(f"\n  📈 YTD Summary:")
    print(f"     Revenue:    ${ytd_revenue:>10,.2f}")
    print(f"     Occ Rate:   {ytd_occ:>9.1%}")
    print(f"     ADR:        ${ytd_adr:>10.2f}")
    print(f"     RevPAR:     ${ytd_revpar:>10.2f}")
    print(f"\n  📅 90-Day Pace (nights on books):")
    print(f"     May: {may.get('nights_sold',0):>3} nights  "
          f"({may.get('occupancy_pct',0):.1%} paced)  "
          f"ADR ${may.get('adr',0):.2f}")
    print(f"     Jun: {jun.get('nights_sold',0):>3} nights  "
          f"({jun.get('occupancy_pct',0):.1%} paced)  "
          f"ADR ${jun.get('adr',0):.2f}")
    print(f"     Jul: {jul.get('nights_sold',0):>3} nights  "
          f"({jul.get('occupancy_pct',0):.1%} paced)  "
          f"ADR ${jul.get('adr',0):.2f}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5: Call Claude API for AI recommendations
# ─────────────────────────────────────────────────────────────────────────────
def get_ai_recommendations(ytd_data, pace_data, as_of_date):
    """
    Calls Claude API with current pace and actuals context.
    Returns 4 recommendation HTML blocks ready to inject into the dashboard.
    Falls back to placeholder text if API call fails.
    """
    import urllib.request
    import json
    import os

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        print("  ⚠️  ANTHROPIC_API_KEY not set — using placeholder recommendations")
        return None

    # Build context from real data
    ytd_revenue     = sum(d["revenue"] for d in ytd_data.values())
    ytd_nights      = sum(d["nights_sold"] for d in ytd_data.values())
    ytd_avail       = sum(d["available_nights"] for d in ytd_data.values())
    ytd_occ         = ytd_nights / ytd_avail if ytd_avail else 0
    ytd_adr         = ytd_revenue / ytd_nights if ytd_nights else 0

    budget_occ = {
        5:0.30, 6:0.30, 7:0.40, 8:0.55, 9:0.20, 10:0.50, 11:0.20, 12:0.10
    }

    pace_lines = []
    for month, data in sorted(pace_data.items()):
        bud = budget_occ.get(month, 0.25)
        occ = data["occupancy_pct"]
        vs  = occ - bud
        pace_lines.append(
            f"  {MONTH_NAMES[month]}: {data['nights_sold']} nights on books "
            f"({occ:.1%} paced vs {bud:.0%} budget, {vs:+.1%} variance)"
            + (f", ADR on books ${data['adr']:.0f}" if data['adr'] > 0 else "")
        )

    prompt = f"""You are an expert hospitality asset manager providing weekly pricing and demand recommendations for a glamping property.

PROPERTY: Ferncrest Flint Creek, Colcord, Oklahoma (NE Oklahoma, serves Tulsa/OKC/NWA/DFW drive market)
UNITS: 12 glamping dome tents (mix of creek-side premium and commons units)
AS-OF DATE: {as_of_date}
BUDGET ADR: $233

YTD ACTUALS (Jan–Apr 2026):
  Revenue: ${ytd_revenue:,.0f}
  Occupancy: {ytd_occ:.1%}
  ADR: ${ytd_adr:.0f}
  Nights sold: {ytd_nights} of {ytd_avail} available

FORWARD PACE (nights currently on books):
{chr(10).join(pace_lines)}

KEY CONTEXT:
  - 40% of bookings historically arrive within 7 days of check-in (last-minute dominant market)
  - Early bookers are least price-sensitive — ADR on books skews high vs final blended rate
  - Creek-side units average $241 ADR vs Commons $193 — rate ladder matters
  - July has only 10 nights but at $395 ADR — strong early-booker signal
  - Oklahoma schools release ~May 22, expect summer bookings to accelerate

Generate exactly 4 pricing and demand recommendations for this week. Each should be specific, actionable, and tied to the actual data above. Consider upcoming holidays, events, and demand patterns for NE Oklahoma.

Respond ONLY with a JSON array of exactly 4 objects, no markdown, no preamble:
[
  {{
    "priority": "urgent|opportunity|monitor",
    "priority_label": "short label (e.g. 'Urgent — 22 days out')",
    "title": "recommendation title (max 8 words)",
    "body": "2-3 sentence explanation tied to specific data points",
    "action": "one concrete action to take this week"
  }}
]"""

    try:
        payload = json.dumps({
            "model": "claude-sonnet-4-6",
            "max_tokens": 1000,
            "messages": [{"role": "user", "content": prompt}]
        }).encode()

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01"
            }
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read())
            recs_json = result["content"][0]["text"].strip()
            # Strip markdown fences if present
            recs_json = recs_json.replace("```json","").replace("```","").strip()
            recs = json.loads(recs_json)
            print(f"  ✅ Got {len(recs)} AI recommendations from Claude API")
            return recs

    except Exception as e:
        print(f"  ⚠️  Claude API call failed: {e} — using placeholder recommendations")
        return None


def build_ai_rec_html(recs):
    """Converts recommendation dicts to HTML blocks for injection."""
    if not recs:
        return None  # Use hardcoded placeholder in template

    html_blocks = []
    priority_labels = {
        "urgent":      ("urgent",      "⚡"),
        "opportunity": ("opportunity", "↑"),
        "monitor":     ("monitor",     "◼"),
    }

    for rec in recs:
        p = rec.get("priority", "monitor")
        css_class, icon = priority_labels.get(p, ("monitor", "◼"))
        label = rec.get("priority_label", p.title())
        title = rec.get("title", "")
        body  = rec.get("body", "")
        action= rec.get("action", "")

        html_blocks.append(f'''      <div class="ai-rec">
        <div class="ai-rec-priority {css_class}">{icon} {label}</div>
        <div class="ai-rec-title">{title}</div>
        <div class="ai-rec-body">{body}</div>
        <div class="ai-rec-action">→ Action: {action}</div>
      </div>''')

    return "\n".join(html_blocks)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6: Inject data into HTML dashboard
# ─────────────────────────────────────────────────────────────────────────────
def generate_html_report(ytd_data, pace_data, as_of_date,
                         template_path="ferncrest_internal_dashboard.html",
                         output_path="ferncrest_internal_dashboard.html"):
    """
    Reads the HTML template, replaces all INJECT placeholders with
    real data, writes the output file.
    """
    import re

    print(f"\n📄 Generating HTML report...")

    if not os.path.exists(template_path):
        print(f"  ⚠️  Template not found: {template_path}")
        return

    with open(template_path, "r") as f:
        html = f.read()

    # ── Compute values ──
    ytd_revenue = sum(d["revenue"] for d in ytd_data.values())
    ytd_nights  = sum(d["nights_sold"] for d in ytd_data.values())
    ytd_avail   = sum(d["available_nights"] for d in ytd_data.values())
    ytd_occ     = ytd_nights / ytd_avail if ytd_avail else 0
    ytd_adr     = ytd_revenue / ytd_nights if ytd_nights else 0
    ytd_revpar  = ytd_revenue / ytd_avail if ytd_avail else 0

    # Budget assumptions (Jan–Apr)
    budget_occ_ytd = {1:0.20, 2:0.20, 3:0.20, 4:0.20}
    budget_adr     = 233
    units          = CONFIG["total_units"]
    days_map       = DAYS_IN_MONTH

    budget_rev_ytd = sum(
        budget_adr * units * budget_occ_ytd.get(m, 0.20) * days_map[m]
        for m in ytd_data.keys()
    )
    budget_occ_avg_ytd = sum(budget_occ_ytd.get(m, 0.20)
                             for m in ytd_data.keys()) / len(ytd_data)
    budget_revpar_ytd  = budget_adr * budget_occ_avg_ytd

    rev_vs_budget    = ytd_revenue - budget_rev_ytd
    occ_vs_budget    = ytd_occ - budget_occ_avg_ytd
    adr_vs_budget    = ytd_adr - budget_adr
    revpar_vs_budget = ytd_revpar - budget_revpar_ytd

    def fmt_dollar(v, show_sign=False):
        sign = "+" if v >= 0 else "−"
        amt  = abs(v)
        if amt >= 1000:
            return f"{'+'if show_sign and v>=0 else sign if v<0 else ''}${amt/1000:.1f}K"
        return f"{'+'if show_sign and v>=0 else sign if v<0 else ''}${amt:.0f}"

    def fmt_pct(v, show_sign=False):
        sign = "+" if v >= 0 else "−"
        return f"{'+' if show_sign and v>=0 else sign if v<0 else ''}{abs(v)*100:.1f} pts"

    # ── Last closed month ──
    last_month = max(ytd_data.keys())
    today = date.today()
    if last_month == today.month and CONFIG["fiscal_year"] == today.year:
        data_through = today.strftime("%B %-d, %Y")
    elif last_month in [4, 6, 9, 11]:
        data_through = f"{MONTH_NAMES[last_month]} 30, {CONFIG['fiscal_year']}"
    else:
        data_through = f"{MONTH_NAMES[last_month]} {days_map[last_month]}, {CONFIG['fiscal_year']}"

    # ── Pace JS data block ──
    budget_occ_fwd = {
        5:0.30, 6:0.30, 7:0.40, 8:0.55, 9:0.20, 10:0.50, 11:0.20, 12:0.10
    }
    today = date.today()
    pace_js_rows = []
    for month in sorted(pace_data.keys()):
        data    = pace_data[month]
        avail   = data["available_nights"]
        on_books= data["nights_sold"]
        bud_occ = budget_occ_fwd.get(month, 0.25)
        adr     = data["adr"]
        arr_date= date(CONFIG["fiscal_year"], month, 1)
        weeks_out = max(0, (arr_date - today).days // 7)
        fy = CONFIG["fiscal_year"]
        mo_name = MONTH_NAMES[month]
        pace_js_rows.append(
            f"  {{month:'{mo_name} {fy}', "
            f"avail:{avail}, onBooks:{on_books}, "
            f"budgetOcc:{bud_occ}, adr:{adr:.2f}, weeksOut:{weeks_out}}},"
        )

    pace_js = "const paceData=[\n" + "\n".join(pace_js_rows) + "\n];"

    # ── AI recommendations ──
    print("  🤖 Fetching AI recommendations...")
    recs     = get_ai_recommendations(ytd_data, pace_data, as_of_date)
    recs_html= build_ai_rec_html(recs)

    # ── Simple placeholder replacer ──
    def inject(html, key, value):
        pattern = f"<!-- INJECT: {key} -->.*?<!-- /INJECT -->"
        replacement = f"<!-- INJECT: {key} -->{value}<!-- /INJECT -->"
        return re.sub(pattern, replacement, html, flags=re.DOTALL)

    # Inject scalar values
    html = inject(html, "generated_date", today.strftime("%B %-d, %Y"))
    html = inject(html, "data_through",   data_through)
    html = inject(html, "week_of",        today.strftime("%B %-d, %Y"))
    html = inject(html, "pace_as_of",     today.strftime("%B %-d, %Y"))

    # KPI values
    html = inject(html, "ytd_revenue",         f"${ytd_revenue/1000:.1f}K")
    html = inject(html, "ytd_rev_vs_budget",   fmt_dollar(rev_vs_budget))
    html = inject(html, "ytd_rev_vs_forecast", fmt_dollar(rev_vs_budget))
    html = inject(html, "ytd_occ",             f"{ytd_occ:.1%}")
    html = inject(html, "ytd_occ_vs_budget",   fmt_pct(occ_vs_budget))
    html = inject(html, "ytd_adr",             f"${ytd_adr:.0f}")
    html = inject(html, "ytd_adr_vs_budget",   fmt_dollar(adr_vs_budget, show_sign=True))
    html = inject(html, "ytd_revpar",          f"${ytd_revpar:.0f}")
    html = inject(html, "ytd_revpar_vs_budget",fmt_dollar(revpar_vs_budget))

    # Pace JS data block
    html = re.sub(
        r"// INJECT: pace_data_js\nconst paceData=\[.*?\];",
        f"// INJECT: pace_data_js\n{pace_js}",
        html, flags=re.DOTALL
    )

    # ── FVA JS data block ──
    budget_occ_all = {
        1:0.20, 2:0.20, 3:0.20, 4:0.20,
        5:0.30, 6:0.30, 7:0.40, 8:0.55, 9:0.20, 10:0.50, 11:0.20, 12:0.10
    }
    budget_adr_val = 233
    fva_rows = []

    # Closed months
    for month, data in sorted(ytd_data.items()):
        if month < today.month:
            bud_rev = round(budget_adr_val * units * budget_occ_all.get(month, 0.20) * days_map[month])
            fva_rows.append(
                f"  {{month:'{MONTH_NAMES[month]}',status:'closed',"
                f"budget:{bud_rev},actual:{round(data['revenue'])},"
                f"budgetOcc:{budget_occ_all.get(month,0.20)},actOcc:{data['occupancy_pct']},"
                f"budgetAdr:{budget_adr_val},actAdr:{data['adr']}}},"
            )

    # Current month
    if today.month in ytd_data or today.month in pace_data:
        m = today.month
        ytd_row  = ytd_data.get(m,  {"revenue":0,"nights_sold":0,"occupancy_pct":0,"adr":0})
        pace_row = pace_data.get(m, {"nights_sold":0,"adr":0,"available_nights": units*days_map.get(m,30)})
        bud_rev  = round(budget_adr_val * units * budget_occ_all.get(m, 0.30) * days_map.get(m, 30))
        bud_nights = round(units * budget_occ_all.get(m, 0.30) * days_map.get(m, 30))
        projected  = round(ytd_row["revenue"] + pace_row["nights_sold"] * (pace_row["adr"] or budget_adr_val))
        pct_paced  = pace_row["nights_sold"] / bud_nights if bud_nights > 0 else 0
        fva_rows.append(
            f"  {{month:'{MONTH_NAMES[m]}',status:'current',"
            f"budget:{bud_rev},actual:{round(ytd_row['revenue'])},projected:{projected},"
            f"budgetOcc:{budget_occ_all.get(m,0.30)},actOcc:{ytd_row['occupancy_pct']},"
            f"pctPaced:{round(pct_paced,3)},budgetNights:{bud_nights},"
            f"onBooks:{pace_row['nights_sold']},adrOnBooks:{pace_row['adr']}}},"
        )

    # Forward months
    for month, data in sorted(pace_data.items()):
        if month > today.month:
            bud_rev    = round(budget_adr_val * units * budget_occ_all.get(month, 0.25) * days_map.get(month, 30))
            bud_nights = round(units * budget_occ_all.get(month, 0.25) * days_map.get(month, 30))
            pct_paced  = data["nights_sold"] / bud_nights if bud_nights > 0 else 0
            fva_rows.append(
                f"  {{month:'{MONTH_NAMES[month]}',status:'forward',"
                f"budget:{bud_rev},budgetOcc:{budget_occ_all.get(month,0.25)},"
                f"actOcc:{data['occupancy_pct']},"
                f"onBooks:{data['nights_sold']},budgetNights:{bud_nights},"
                f"pctPaced:{round(pct_paced,3)},adrOnBooks:{data['adr']}}},"
            )

    fva_js = "const fvaData=[\n" + "\n".join(fva_rows) + "\n];"
    html = re.sub(
        r"// INJECT: fva_data_js\nconst fvaData=\[.*?\];",
        f"// INJECT: fva_data_js\n{fva_js}",
        html, flags=re.DOTALL
    )

    # AI recommendations (only if API returned something)
    if recs_html:
        html = re.sub(
            r"<!-- INJECT: ai_recommendations -->.*?<!-- /INJECT -->",
            f"<!-- INJECT: ai_recommendations -->\n{recs_html}\n      <!-- /INJECT -->",
            html, flags=re.DOTALL
        )

    # Fix KPI variance CSS classes based on actual direction
    # Revenue behind → rust, ADR ahead → sage
    for key, val in [
        ("ytd_rev_vs_budget",    rev_vs_budget >= 0),
        ("ytd_rev_vs_forecast",  rev_vs_budget >= 0),
        ("ytd_occ_vs_budget",    occ_vs_budget >= 0),
        ("ytd_adr_vs_budget",    adr_vs_budget >= 0),
        ("ytd_revpar_vs_budget", revpar_vs_budget >= 0),
    ]:
        css_class = "ahead" if val else "behind"
        html = html.replace(
            f'<!-- INJECT: {key} -->',
            f'<!-- INJECT: {key} -->'
        )

    with open(output_path, "w") as f:
        f.write(html)

    # Also save to Google Drive folder
    gdrive_path = os.path.join(_GDRIVE, "ferncrest_internal_dashboard.html")
    try:
        with open(gdrive_path, "w") as f:
            f.write(html)
        print(f"  ✅ HTML report saved: {output_path} + Google Drive")
    except Exception:
        print(f"  ✅ HTML report saved: {output_path} (Google Drive save failed)")

    print(f"     Revenue: ${ytd_revenue:,.0f} | Occ: {ytd_occ:.1%} | ADR: ${ytd_adr:.0f} | RevPAR: ${ytd_revpar:.0f}")

    # Auto-push updated HTML to FlintCreekReport GitHub Pages repo
    import subprocess
    flint_repo = os.path.expanduser("~/FlintCreekReport")
    if os.path.isdir(flint_repo):
        try:
            import shutil
            shutil.copy(output_path, os.path.join(flint_repo, "index.html"))
            subprocess.run(["git", "-C", flint_repo, "add", "index.html"], check=True)
            subprocess.run(["git", "-C", flint_repo, "commit", "-m",
                            f"Update report {as_of_date}"], check=True)
            subprocess.run(["git", "-C", flint_repo, "push"], check=True)
            print(f"  ✅ Live report updated: https://linton-hospitality.github.io/FlintCreekReport/")
        except Exception as e:
            print(f"  ⚠️  Live report push failed: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7: Post Slack digest
# ─────────────────────────────────────────────────────────────────────────────
def _flatten_daily_rows(parsed_data):
    """Combines daily_rows across all months of a parsed Cloudbeds export
    into a single {MM/DD: row} lookup."""
    out = {}
    for data in parsed_data.values():
        for row in data.get("daily_rows", []):
            out[row["date"]] = row
    return out


def _load_pace_snapshot(path):
    """Loads last week's daily on-books snapshot, if one was saved."""
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return None
    return None


def _save_pace_snapshot(path, pace_daily, as_of_date):
    """Saves this week's daily on-books + revenue snapshot for next week's
    pickup comparison."""
    snapshot = {
        "as_of_date": as_of_date,
        "on_books":   {d: row["nights_sold"] for d, row in pace_daily.items()},
        "revenue":    {d: row["revenue"]     for d, row in pace_daily.items()},
    }
    try:
        with open(path, "w") as f:
            json.dump(snapshot, f, indent=2)
    except OSError as e:
        print(f"  ⚠️  Could not save pace snapshot cache: {e}")


def _compute_pickup_stats(pace_daily, prev_snapshot, today, fiscal_year):
    """Compares this week's daily on-books snapshot to last week's to derive
    net pickup, ADR on new bookings, and pickup for near-term arrivals.
    Returns None if no prior snapshot exists yet (first run)."""
    if not prev_snapshot:
        return None

    prev_books = prev_snapshot.get("on_books", {})
    prev_rev   = prev_snapshot.get("revenue", {})

    pickup_total      = 0
    pickup_14d        = 0
    rev_delta         = 0.0
    nights_delta      = 0
    net_revenue_pickup = 0.0

    for d, row in pace_daily.items():
        delta = row["nights_sold"] - prev_books.get(d, 0)
        pickup_total += delta
        net_revenue_pickup += row["revenue"] - prev_rev.get(d, 0)

        if delta > 0:
            rev_delta    += row["revenue"] - prev_rev.get(d, 0)
            nights_delta += delta

            try:
                m, dd = map(int, d.split("/"))
                row_date = date(fiscal_year, m, dd)
                if 0 <= (row_date - today).days <= 14:
                    pickup_14d += delta
            except (ValueError, AttributeError):
                pass

    adr_new = (rev_delta / nights_delta) if nights_delta > 0 else 0

    return {
        "pickup_7d":           pickup_total,
        "pickup_14d":          pickup_14d,
        "adr_new":             adr_new,
        "revenue_pickup":      net_revenue_pickup,
    }


def _build_forward_pace_table(pace_daily, today, weeks=5):
    """Builds the 'next N weeks' forward pace table against OK-FC pace targets."""
    units = CONFIG["total_units"]
    monday_this_week = today - timedelta(days=today.weekday())

    rows = []
    for week_offset in range(weeks):
        week_start = monday_this_week + timedelta(weeks=week_offset)
        weeks_out  = week_offset + 1

        booked = 0
        for i in range(7):
            key = (week_start + timedelta(days=i)).strftime("%m/%d")
            row = pace_daily.get(key)
            if row:
                booked += row["nights_sold"]

        avail      = units * 7
        booked_pct = booked / avail if avail else 0
        month      = week_start.month

        monthly_target = TARGET_OCC_OKFC.get(month, 0.25)
        pace_factor    = PACE_TARGETS_OKFC.get(weeks_out, PACE_TARGETS_OKFC[5])
        target_pct     = monthly_target * (pace_factor / PACE_TARGETS_OKFC[1])

        if weeks_out == 5:
            flag = "⬜"
        elif weeks_out == 4:
            flag = "🟡" if booked_pct < target_pct * 0.80 else "✅"
        elif booked_pct >= target_pct * 1.05:
            flag = "✅"
        elif booked_pct >= target_pct * 0.80:
            flag = "🟡"
        else:
            flag = "🔴"

        rows.append({
            "week_label": f"{week_start.strftime('%b %-d')} wk",
            "booked":     booked,
            "avail":      avail,
            "booked_pct": booked_pct,
            "target_pct": target_pct,
            "weeks_out":  weeks_out,
            "flag":       flag,
        })

    return rows


def _build_weekend_watch(pace_daily, today, weekends=4):
    """Next N weekends (Fri+Sat), flagged per OK-FC weekend thresholds."""
    units = CONFIG["total_units"]
    days_until_friday = (4 - today.weekday()) % 7
    next_friday = today + timedelta(days=days_until_friday)

    rows = []
    for i in range(weekends):
        fri = next_friday + timedelta(weeks=i)
        sat = fri + timedelta(days=1)

        fri_row = pace_daily.get(fri.strftime("%m/%d"))
        sat_row = pace_daily.get(sat.strftime("%m/%d"))
        fri_booked = fri_row["nights_sold"] if fri_row else 0
        sat_booked = sat_row["nights_sold"] if sat_row else 0

        fri_pct = fri_booked / units
        sat_pct = sat_booked / units
        avg_pct = (fri_booked + sat_booked) / (units * 2)

        if i == 0 and (fri_pct < 0.5 or sat_pct < 0.5):
            flag = "🔴"
        elif fri_pct < 0.30 and i <= 2:
            flag = "🔴"
        elif fri_pct < 0.60 and i <= 1:
            flag = "🟡"
        else:
            flag = "🟢"

        rows.append({
            "label":      f"{fri.strftime('%b %-d')} wknd",
            "weeks_out":  i,
            "fri_booked": fri_booked,
            "sat_booked": sat_booked,
            "avg_pct":    avg_pct,
            "flag":       flag,
        })

    return rows


def _get_action_item(pickup_stats, pace_rows, weekend_rows):
    """Calls Claude for a single-sentence action item; falls back to a
    heuristic if the API key isn't set or the call fails."""
    import urllib.request

    # Worst pace gap in the action zone (weeks 1-4, excludes the horizon week)
    worst = min(pace_rows[:4], key=lambda r: r["booked_pct"] - r["target_pct"])
    weekend_summary = ", ".join(f"{w['label']} {w['flag']}" for w in weekend_rows)

    if pickup_stats:
        pickup_line = (f"Pickup last 7 days: {pickup_stats['pickup_7d']:+d} nights "
                        f"(ADR on new bookings: ${pickup_stats['adr_new']:.0f}, "
                        f"{pickup_stats['pickup_14d']} for arrivals within 14 days, "
                        f"${pickup_stats['revenue_pickup']:+,.0f} net revenue pickup)")
    else:
        pickup_line = "Pickup last 7 days: no prior snapshot yet — comparison starts next week"

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if api_key:
        prompt = f"""You are the revenue manager for Ferncrest Flint Creek, a 12-key glamping property in Colcord, Oklahoma.
Audience: internal marketing team and CEO. Tone: direct, one sentence, specific.

Data for this week:
- {pickup_line}
- Most urgent pace gap: {worst['week_label']} at {worst['booked_pct']:.0%} booked vs {worst['target_pct']:.0%} target
- Weekend flags: {weekend_summary}

Property context:
- Peak season is Jul-Aug with last-minute booking pattern (median 5-6 days lead time)
- October is the strongest ADR month ($287 in 2025) — fall foliage driver
- Commons units (Tents 6-12) have lowest occupancy — group bookings are the primary lever
- Do not recommend rate cuts — occupancy gap is a demand/awareness problem not a price problem
- Do not recommend discounting weekends

Write one action sentence for the marketing team. Be specific about which week or weekend needs attention
and what action to take (e.g. push organic content, send email to past guests, boost paid spend)."""

        try:
            payload = json.dumps({
                "model": "claude-sonnet-4-6",
                "max_tokens": 200,
                "messages": [{"role": "user", "content": prompt}]
            }).encode()
            req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages",
                data=payload,
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01"
                }
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                result = json.loads(resp.read())
                text = result["content"][0]["text"].strip()
                print("  ✅ AI action item generated")
                return text
        except Exception as e:
            print(f"  ⚠️  Claude API call failed: {e} — using heuristic action item")

    # Heuristic fallback
    if worst["flag"] == "🔴":
        return (f"{worst['week_label']} is the most urgent gap at {worst['booked_pct']:.0%} booked vs "
                f"{worst['target_pct']:.0%} target — push organic content and email past guests this week.")
    urgent_wknd = next((w for w in weekend_rows if w["flag"] == "🔴"), None)
    if urgent_wknd:
        return (f"{urgent_wknd['label']} is under-booked — boost paid spend targeting "
                f"last-minute drive-market guests.")
    return "No urgent pace or weekend gaps this week — hold steady and monitor."


def post_slack_digest(ytd_data, pace_data, as_of_date):
    """
    Posts the weekly Monday digest to Slack: pickup, forward pace,
    weekend watch, and one action item.
    """
    import urllib.request

    webhook_url = os.environ.get("SLACK_WEBHOOK_URL", "")
    if not webhook_url:
        print("\n📣 Slack webhook not configured — skipping post")
        print("   Set SLACK_WEBHOOK_URL environment variable to enable")
        return

    print("\n📣 Building Slack digest...")

    today       = date.today()
    fiscal_year = CONFIG["fiscal_year"]
    units       = CONFIG["total_units"]
    pace_daily  = _flatten_daily_rows(pace_data)

    # ── Pickup vs last week's snapshot ──
    prev_snapshot = _load_pace_snapshot(CONFIG["snapshot_cache"])
    pickup_stats  = _compute_pickup_stats(pace_daily, prev_snapshot, today, fiscal_year)
    _save_pace_snapshot(CONFIG["snapshot_cache"], pace_daily, as_of_date)

    if pickup_stats:
        pickup_text = (
            f"• *{pickup_stats['pickup_7d']:+d} nights* net pickup in last 7 days"
            + (f" · ADR on new bookings: *${pickup_stats['adr_new']:.0f}*"
               if pickup_stats["adr_new"] > 0 else "")
            + f"\n• *{pickup_stats['pickup_14d']} of those* for arrivals within the next 14 days"
            + f"\n• *${pickup_stats['revenue_pickup']:+,.0f}* net revenue picked up vs last week"
        )
    else:
        pickup_text = "• First snapshot recorded — pickup comparison starts next week"

    # ── Forward pace table (next 5 weeks) ──
    pace_rows = _build_forward_pace_table(pace_daily, today)

    # Flag the single worst 🔴 week among weeks 1-2 for "← push now"
    push_now_idx = next(
        (i for i, r in enumerate(pace_rows[:2]) if r["flag"] == "🔴"), None
    )

    pace_lines = []
    for idx, r in enumerate(pace_rows):
        if r["weeks_out"] == 5:
            line = (f"{r['flag']}  {r['week_label']:<10} —  "
                    f"{r['booked']:>2}/{r['avail']} ({r['booked_pct']:.0%})  —")
        else:
            line = (f"{r['flag']}  {r['week_label']:<10} —  "
                    f"{r['booked']:>2}/{r['avail']} ({r['booked_pct']:.0%})  "
                    f"tgt {r['target_pct']:.0%}")
            if idx == push_now_idx:
                line += "  ← push now"
        pace_lines.append(line)
    pace_text = "\n".join(pace_lines)

    # ── Weekend watch (next 4 weekends) ──
    weekend_rows = _build_weekend_watch(pace_daily, today)
    weekend_lines = [
        f"{w['flag']} {w['label']} — {w['weeks_out']}wk{'s' if w['weeks_out'] != 1 else ''} out · "
        f"Fri {w['fri_booked']}/{units} · Sat {w['sat_booked']}/{units} · {w['avg_pct']:.0%} booked"
        for w in weekend_rows
    ]
    weekend_text = "\n".join(weekend_lines)

    # ── Action item ──
    action_text = _get_action_item(pickup_stats, pace_rows, weekend_rows)

    # ── Build Slack blocks ──
    blocks = [
        {
            "type": "header",
            "text": {"type": "plain_text", "text": f"🏕 Flint Creek — Monday Digest · {as_of_date}"}
        },
        {"type": "divider"},
        {
            "type": "section",
            "text": {"type": "mrkdwn", "text": f"📊 *This week's pickup*\n{pickup_text}"}
        },
        {"type": "divider"},
        {
            "type": "section",
            "text": {"type": "mrkdwn", "text": f"📅 *Forward pace — next 5 weeks*\n```{pace_text}```"}
        },
        {"type": "divider"},
        {
            "type": "section",
            "text": {"type": "mrkdwn", "text": f"🏖 *Weekend watch*\n{weekend_text}"}
        },
        {"type": "divider"},
        {
            "type": "section",
            "text": {"type": "mrkdwn", "text": f"🎯 *This week:* {action_text}"}
        },
        {
            "type": "context",
            "elements": [
                {
                    "type": "mrkdwn",
                    "text": (f"<{CONFIG['full_report_url']}|View Full Report>  ·  "
                             f"Ferncrest Pipeline · {as_of_date}")
                }
            ]
        }
    ]

    try:
        payload = json.dumps({"blocks": blocks}).encode()
        req = urllib.request.Request(
            webhook_url,
            data=payload,
            headers={"Content-Type": "application/json"}
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            print(f"  ✅ Slack digest posted (status {resp.status})")
    except Exception as e:
        print(f"  ⚠️  Slack post failed: {e}")
def main():
    print("=" * 60)
    print("  FERNCREST DATA PIPELINE")
    print(f"  Property: {CONFIG['property_code']}")
    print(f"  Run date: {date.today().strftime('%Y-%m-%d')}")
    print("=" * 60)

    # Parse both Cloudbeds exports
    ytd_data  = parse_cloudbeds_export(CONFIG["ytd_file"])
    pace_data = parse_cloudbeds_export(CONFIG["pace_file"])

    # Load the property workbook
    wb_path = CONFIG["workbook"]
    if not os.path.exists(wb_path):
        print(f"\n❌ Workbook not found: {wb_path}")
        print("   Place ferncrest_ok-fc_v2.xlsx in the same folder as this script.")
        sys.exit(1)

    print(f"\n📖 Loading workbook: {wb_path}")
    wb = load_workbook(wb_path)
    print(f"   Sheets: {wb.sheetnames}")

    # Write to tabs
    write_actuals(wb["Actuals_Input"], ytd_data)
    write_pace(wb["Pace_Weekly"], pace_data)
    update_report_data(wb["Report_Data"], ytd_data, pace_data)

    # Save workbook
    wb.save(wb_path)
    print(f"\n💾 Workbook saved: {wb_path}")

    # Generate HTML report
    generate_html_report(
        ytd_data, pace_data, as_of_date=date.today().strftime("%Y-%m-%d"),
        template_path="ferncrest_internal_dashboard.html",
        output_path="ferncrest_internal_dashboard.html"
    )

    # Post Slack digest
    post_slack_digest(ytd_data, pace_data,
                      as_of_date=date.today().strftime("%Y-%m-%d"))

    print("\n✅ Pipeline complete.")
    print("=" * 60)

    return {
        "ytd": ytd_data,
        "pace": pace_data,
    }


if __name__ == "__main__":
    main()


# ─────────────────────────────────────────────────────────────────────────────
