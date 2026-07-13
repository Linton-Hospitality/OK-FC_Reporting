#!/usr/bin/env python3
"""
Ferncrest Promised Land (PA-PL) Data Pipeline
==============================================
Structural copy of ferncrest_pipeline.py (OK-FC), adapted for Promised Land's
24-unit, 7-type inventory with a seasonal unit-availability calendar (unlike
OK-FC's flat 12-unit count) and a sharply month-dependent booking lead-time
pattern (peak summer books 9-11 days out; April/May/Sept book 30-39 days out).

Kept as an independent script per OK-FC's pattern — only worth merging into a
shared module once a third property exists.

File expectations (set paths in CONFIG below):
    - PA-PL_occupancy_2026ytd.xlsx       : Jan 1 onward actuals
    - PA-PL_occupancy_pace_150day.xlsx   : forward 150-day pace
    - pace_snapshot_cache_pl.json        : last week's daily on-books snapshot

What it does:
    1. Parses both Cloudbeds exports
    2. Aggregates daily rows -> monthly totals using a date-based available-
       units calendar (9/15/24 units depending on time of year)
    3. Calculates ADR, occupancy %, RevPAR per month
    4. Builds pickup vs last week, forward pace, weekend watch, and an AI
       action item
    5. Posts the weekly Monday Slack digest
"""

import openpyxl
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime, date, timedelta
import os
import sys
import json
from dotenv import load_dotenv

_GDRIVE = "/Users/stephcleung/Library/CloudStorage/GoogleDrive-stephanie@lintonhospitality.com/Shared drives/Ferncrest/1_Locations/01_PA-PL/6_Reports_PA_PL/AssetManagement"
_CI = os.environ.get("CI") == "true"
_BASE = os.path.dirname(os.path.abspath(__file__)) if _CI else _GDRIVE
_REPO = os.path.dirname(os.path.abspath(__file__))  # always the repo dir, for snapshot cache

load_dotenv(os.path.join(_BASE, ".env"), override=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG — paths auto-switch between local Mac and GitHub Actions
# ─────────────────────────────────────────────────────────────────────────────
CONFIG = {
    "ytd_file":        os.path.join(_BASE, "PA-PL_occupancy_2026ytd.xlsx"),
    "pace_file":       os.path.join(_BASE, "PA-PL_occupancy_pace_150day.xlsx"),
    "py_file":         os.path.join(_REPO, "PA-PL_occupancy_2025.xlsx"),
    "workbook":        os.path.join(_BASE, "ferncrest_pa-pl_v1.xlsx"),
    "snapshot_cache":  os.path.join(_REPO, "pace_snapshot_cache_pl.json"),
    "full_report_url": "https://linton-hospitality.github.io/PromisedLandReport/",  # TBD — confirm repo
    "property_code":   "PA-PL",
    "fiscal_year":     2026,
}

# ── Seasonal available-units calendar ──────────────────────────────────────
# Unlike OK-FC's flat 12 units, PA-PL's inventory opens/closes by date:
#   Jan-Mar:            9  (winter cohort only)
#   Apr 1-2:             9
#   Apr 3 - May 20:     15  (+ standard domes)
#   May 21 - Oct 27:    24  (full inventory, crest tents open)
#   Oct 28 - Nov 2:     15  (crest tents closed)
#   Nov 3 - Nov 30:      9  (standard domes closed)
#   Dec:                 9
def units_available_on(d: date) -> int:
    if d.month in (1, 2, 3):
        return 9
    if d.month == 4:
        return 9 if d.day < 3 else 15
    if d.month == 5:
        return 15 if d.day < 21 else 24
    if d.month in (6, 7, 8, 9):
        return 24
    if d.month == 10:
        return 24 if d.day <= 27 else 15
    if d.month == 11:
        return 15 if d.day <= 2 else 9
    return 9  # December


def avail_nights_for_month(year: int, month: int) -> int:
    """Sums daily available units across every day in the given month."""
    days = DAYS_IN_MONTH.get(month, 30)
    return sum(units_available_on(date(year, month, d)) for d in range(1, days + 1))


# Target monthly occupancy — ESTIMATED from the narrative seasonality curve
# (Aug 50%+ peak, Jul/Oct strong secondary, Jun ramp, Apr/May soft-open,
# Sep genuine-but-not-peak demand, Nov/Dec weekend-driven winter).
# These are placeholders pending your confirmation/adjustment — unlike
# OK-FC, no specific numeric targets were given for PA-PL.
TARGET_OCC_PROMISEDLAND = {
    1: 0.15, 2: 0.15, 3: 0.15, 4: 0.10,
    5: 0.20, 6: 0.35, 7: 0.45, 8: 0.55,
    9: 0.30, 10: 0.45, 11: 0.20, 12: 0.20,
}

# Pace targets by weeks-out — ESTIMATED, and deliberately NOT a single curve
# like OK-FC's PACE_TARGETS_OKFC. PA-PL's booking lead time varies sharply by
# month (Jul/Aug median 9-11 days; April 37 days; Sept 39 days), so a single
# weeks-out curve would misfire — e.g. flagging July as "behind" in May/June
# when thin summer pace that far out is structurally normal.
# "last_minute" months: thin pace until ~1-2 weeks out is normal, no flag.
# "early_booking" months: pace should build steadily from ~5 weeks out.
PACE_PROFILE_BY_MONTH = {
    1: "winter_short", 2: "winter_short", 3: "early_booking",
    4: "early_booking", 5: "early_booking",
    6: "last_minute", 7: "last_minute", 8: "last_minute",
    9: "early_booking", 10: "moderate",
    11: "winter_short", 12: "winter_short",
}

PACE_TARGETS_BY_PROFILE = {
    # weeks_out: fraction of monthly target occupancy expected booked by then
    "last_minute":    {5: 0.05, 4: 0.10, 3: 0.15, 2: 0.25, 1: 0.45},
    "early_booking":  {5: 0.15, 4: 0.30, 3: 0.45, 2: 0.60, 1: 0.75},
    "moderate":       {5: 0.10, 4: 0.20, 3: 0.35, 2: 0.50, 1: 0.65},
    "winter_short":   {5: 0.08, 4: 0.15, 3: 0.25, 2: 0.40, 1: 0.60},
}

DAYS_IN_MONTH = {
    1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30,
    7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31
}

MONTH_NAMES = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
}

# Known anomalies to exclude from rate benchmarking (informational — not
# auto-excluded by this script since current exports don't carry a year-prior
# column; flag here so a human reviewing output knows the context).
RATE_ANOMALIES = {
    "2025-09": "RM pricing error compressed ADR to ~$171; true demand ~47% occ. Use 2024-09 as rate reference.",
    "2026-03": "BFCM 40%-off promo redeemed in March; $179 ADR is real but not a forward pricing benchmark.",
}


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: Parse a Cloudbeds occupancy export
# ─────────────────────────────────────────────────────────────────────────────
def parse_cloudbeds_export(filepath):
    """
    Reads a Cloudbeds occupancy xlsx export (Date, ADR <yr>, ACCOMMODATIONS
    BOOKED <yr>, Room Rates <yr>). Returns a dict keyed by month number:
        {month: {nights_sold, available_nights, revenue, occupancy_pct,
                  adr, revpar, daily_rows}}
    available_nights is computed from the date-based units calendar, not a
    flat unit count.
    """
    print(f"\n📂 Reading: {filepath}")

    if not os.path.exists(filepath):
        print(f"  ❌ File not found: {filepath}")
        sys.exit(1)

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    headers = rows[0]
    expected_prefixes = ("Date", "ADR", "ACCOMMODATIONS BOOKED", "Room Rates")
    if not all(str(h).startswith(p) for h, p in zip(headers, expected_prefixes)):
        print(f"  ⚠️  Unexpected headers: {headers}")
        print(f"       Expected prefixes: {expected_prefixes}")
        print(f"       Column mapping may be off — check before trusting output.")

    fiscal_year = CONFIG["fiscal_year"]
    monthly = defaultdict(lambda: {
        "nights_sold": 0,
        "revenue": 0.0,
        "available_nights": 0,
        "daily_rows": []
    })

    skipped = 0
    for row in rows[1:]:
        date_str, adr_val, booked, revenue = row

        if date_str is None:
            skipped += 1
            continue

        try:
            month, day = (int(p) for p in str(date_str).split("/"))
        except (ValueError, IndexError):
            skipped += 1
            continue

        nights = int(booked) if booked else 0
        rev = float(revenue) if revenue else 0.0

        monthly[month]["nights_sold"] += nights
        monthly[month]["revenue"] += rev
        monthly[month]["available_nights"] += units_available_on(date(fiscal_year, month, day))
        monthly[month]["daily_rows"].append({
            "date": date_str,
            "adr": float(adr_val) if adr_val else 0.0,
            "nights_sold": nights,
            "revenue": rev,
        })

    if skipped:
        print(f"  ℹ️  Skipped {skipped} blank rows")

    result = {}
    for month, data in sorted(monthly.items()):
        avail = data["available_nights"]
        nights_sold = data["nights_sold"]
        revenue = data["revenue"]
        occ = nights_sold / avail if avail > 0 else 0
        adr = revenue / nights_sold if nights_sold > 0 else 0
        revpar = revenue / avail if avail > 0 else 0

        result[month] = {
            "nights_sold":      nights_sold,
            "available_nights": avail,
            "revenue":          round(revenue, 2),
            "occupancy_pct":    round(occ, 4),
            "adr":              round(adr, 2),
            "revpar":           round(revpar, 2),
            "daily_rows":       data["daily_rows"],
        }

        print(f"  {MONTH_NAMES[month]:>3}: "
              f"nights={nights_sold:>3} / {avail}  "
              f"occ={occ:>5.1%}  "
              f"ADR=${adr:>7.2f}  "
              f"rev=${revenue:>9,.2f}")

    return result


def parse_prior_year_daily(filepath):
    """
    Reads a full-year Cloudbeds export (e.g. PA-PL_occupancy_2025.xlsx) and
    returns {"MM/DD": {"adr", "nights_sold", "revenue"}} for same-day-of-year
    YoY comparison. No occupancy % here — PY's unit count ramped on different
    dates than the current year's calendar (units_available_on), and exact PY
    transition dates haven't been confirmed, so PY occupancy/RevPAR would be
    a guess. ADR and nights-sold are denominator-free and safe to compare.
    """
    if not os.path.exists(filepath):
        print(f"  ℹ️  No prior-year file at {filepath} — skipping YoY comparison")
        return {}

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    out = {}
    for row in rows[1:]:
        date_str, adr_val, booked, revenue = row
        if date_str is None:
            continue
        out[str(date_str)] = {
            "adr":         float(adr_val) if adr_val else 0.0,
            "nights_sold": int(booked) if booked else 0,
            "revenue":     float(revenue) if revenue else 0.0,
        }
    return out


def _build_yoy_week_comparison(ytd_daily_all, py_daily, today, weeks=4):
    """Compares the last N fully-closed weeks of 2026 actuals vs same calendar
    weeks in 2025. Uses ytd_daily_all (closed actuals) — not forward pace — so
    both sides are final realized numbers and the comparison is apples-to-apples."""
    # Last fully-closed week ended last Sunday
    last_sunday = today - timedelta(days=today.weekday() + 1)
    rows = []
    for week_offset in range(weeks):
        week_end = last_sunday - timedelta(weeks=week_offset)
        week_start = week_end - timedelta(days=6)
        cur_nights, cur_rev = 0, 0.0
        py_nights, py_rev = 0, 0.0
        anomaly_note = None
        for i in range(7):
            day = week_start + timedelta(days=i)
            key = day.strftime("%m/%d")
            cur_row = ytd_daily_all.get(key)
            if cur_row:
                cur_nights += cur_row["nights_sold"]
                cur_rev += cur_row["revenue"]
            py_row = py_daily.get(key)
            if py_row:
                py_nights += py_row["nights_sold"]
                py_rev += py_row["revenue"]
            py_key = f"2025-{day.month:02d}"
            if py_key in RATE_ANOMALIES:
                anomaly_note = RATE_ANOMALIES[py_key]

        cur_adr = cur_rev / cur_nights if cur_nights else 0
        py_adr = py_rev / py_nights if py_nights else 0
        rows.append({
            "week_label":  f"{week_start.strftime('%b %-d')} wk",
            "cur_nights":  cur_nights,
            "py_nights":   py_nights,
            "cur_adr":     cur_adr,
            "py_adr":      py_adr,
            "nights_yoy":  (cur_nights - py_nights),
            "adr_yoy_pct": ((cur_adr - py_adr) / py_adr) if py_adr else None,
            "anomaly":     anomaly_note,
        })
    rows.reverse()  # chronological order, most recent last
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: Pickup vs last week's snapshot
# ─────────────────────────────────────────────────────────────────────────────
def _flatten_daily_rows(parsed_data):
    out = {}
    for data in parsed_data.values():
        for row in data.get("daily_rows", []):
            out[row["date"]] = row
    return out


def _load_pace_snapshot(path):
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return None
    return None


def _save_pace_snapshot(path, pace_daily, as_of_date):
    snapshot = {
        "as_of_date": as_of_date,
        "on_books":   {d: row["nights_sold"] for d, row in pace_daily.items()},
        "revenue":    {d: row["revenue"]     for d, row in pace_daily.items()},
    }
    try:
        with open(path, "w") as f:
            json.dump(snapshot, f, indent=2)
    except OSError as e:
        print(f"  ⚠️  Could not save pace snapshot cache: {e}")


def _compute_pickup_stats(pace_daily, prev_snapshot, today, fiscal_year):
    if not prev_snapshot:
        return None

    prev_books = prev_snapshot.get("on_books", {})
    prev_rev = prev_snapshot.get("revenue", {})

    pickup_total = 0
    pickup_14d = 0
    rev_delta = 0.0
    nights_delta = 0
    net_revenue_pickup = 0.0

    for d, row in pace_daily.items():
        delta = row["nights_sold"] - prev_books.get(d, 0)
        pickup_total += delta
        net_revenue_pickup += row["revenue"] - prev_rev.get(d, 0)

        if delta > 0:
            rev_delta += row["revenue"] - prev_rev.get(d, 0)
            nights_delta += delta
            try:
                m, dd = map(int, d.split("/"))
                row_date = date(fiscal_year, m, dd)
                if 0 <= (row_date - today).days <= 14:
                    pickup_14d += delta
            except (ValueError, AttributeError):
                pass

    adr_new = (rev_delta / nights_delta) if nights_delta > 0 else 0

    return {
        "pickup_7d":      pickup_total,
        "pickup_14d":     pickup_14d,
        "adr_new":        adr_new,
        "revenue_pickup": net_revenue_pickup,
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: Forward pace table — month-aware lead-time profile
# ─────────────────────────────────────────────────────────────────────────────
def _build_forward_pace_table(pace_daily, today, weeks=5):
    """Builds the 'next N weeks' forward pace table. Uses a pace-target curve
    keyed to each week's month-specific booking-lead profile, so summer's
    structurally thin early pace doesn't get mis-flagged as behind."""
    monday_this_week = today - timedelta(days=today.weekday())

    rows = []
    for week_offset in range(weeks):
        week_start = monday_this_week + timedelta(weeks=week_offset)
        weeks_out = week_offset + 1

        booked = 0
        avail = 0
        for i in range(7):
            day = week_start + timedelta(days=i)
            key = day.strftime("%m/%d")
            row = pace_daily.get(key)
            if row:
                booked += row["nights_sold"]
            avail += units_available_on(day)

        booked_pct = booked / avail if avail else 0
        month = week_start.month

        monthly_target = TARGET_OCC_PROMISEDLAND.get(month, 0.25)
        profile = PACE_PROFILE_BY_MONTH.get(month, "moderate")
        curve = PACE_TARGETS_BY_PROFILE[profile]
        pace_factor = curve.get(weeks_out, curve[5])
        target_pct = monthly_target * (pace_factor / curve[1])

        if weeks_out == 5:
            flag = "⬜"
        elif weeks_out == 4:
            flag = "🟡" if booked_pct < target_pct * 0.80 else "✅"
        elif booked_pct >= target_pct * 1.05:
            flag = "✅"
        elif booked_pct >= target_pct * 0.80:
            flag = "🟡"
        else:
            flag = "🔴"

        rows.append({
            "week_label": f"{week_start.strftime('%b %-d')} wk",
            "booked":     booked,
            "avail":      avail,
            "booked_pct": booked_pct,
            "target_pct": target_pct,
            "weeks_out":  weeks_out,
            "profile":    profile,
            "flag":       flag,
        })

    return rows


def _build_weekend_watch(pace_daily, today, weekends=4):
    """Next N weekends (Fri+Sat). Uses the actual units open that weekend
    (not a flat count) since PA-PL's inventory varies seasonally."""
    days_until_friday = (4 - today.weekday()) % 7
    next_friday = today + timedelta(days=days_until_friday)

    rows = []
    for i in range(weekends):
        fri = next_friday + timedelta(weeks=i)
        sat = fri + timedelta(days=1)

        fri_units = units_available_on(fri)
        sat_units = units_available_on(sat)

        fri_row = pace_daily.get(fri.strftime("%m/%d"))
        sat_row = pace_daily.get(sat.strftime("%m/%d"))
        fri_booked = fri_row["nights_sold"] if fri_row else 0
        sat_booked = sat_row["nights_sold"] if sat_row else 0

        fri_pct = fri_booked / fri_units if fri_units else 0
        sat_pct = sat_booked / sat_units if sat_units else 0
        avg_pct = (fri_booked + sat_booked) / (fri_units + sat_units) if (fri_units + sat_units) else 0

        if i == 0 and (fri_pct < 0.5 or sat_pct < 0.5):
            flag = "🔴"
        elif fri_pct < 0.30 and i <= 2:
            flag = "🔴"
        elif fri_pct < 0.60 and i <= 1:
            flag = "🟡"
        else:
            flag = "🟢"

        rows.append({
            "label":      f"{fri.strftime('%b %-d')} wknd",
            "weeks_out":  i,
            "fri_booked": fri_booked,
            "sat_booked": sat_booked,
            "fri_units":  fri_units,
            "sat_units":  sat_units,
            "avg_pct":    avg_pct,
            "flag":       flag,
        })

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: AI action item
# ─────────────────────────────────────────────────────────────────────────────
def _get_action_item(pickup_stats, pace_rows, weekend_rows):
    """Calls Claude for a single-sentence action item; falls back to a
    heuristic if the API key isn't set or the call fails."""
    import urllib.request

    worst = min(pace_rows[:4], key=lambda r: r["booked_pct"] - r["target_pct"])
    weekend_summary = ", ".join(f"{w['label']} {w['flag']}" for w in weekend_rows)

    if pickup_stats:
        pickup_line = (f"Pickup last 7 days: {pickup_stats['pickup_7d']:+d} nights "
                        f"(ADR on new bookings: ${pickup_stats['adr_new']:.0f}, "
                        f"{pickup_stats['pickup_14d']} for arrivals within 14 days, "
                        f"${pickup_stats['revenue_pickup']:+,.0f} net revenue pickup)")
    else:
        pickup_line = "Pickup last 7 days: no prior snapshot yet — comparison starts next week"

    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if api_key:
        prompt = f"""You are the revenue manager for Ferncrest Promised Land, a 24-unit glamping property in Pennsylvania serving the NYC/Philadelphia drive market.
Audience: internal marketing team and CEO. Tone: direct, one sentence, specific.

Data for this week:
- {pickup_line}
- Most urgent pace gap: {worst['week_label']} ({worst['profile']} lead-time profile) at {worst['booked_pct']:.0%} booked vs {worst['target_pct']:.0%} target
- Weekend flags: {weekend_summary}

Property context:
- Peak season is August (50%+ occ when correctly priced); July and October are strong secondary months
- CRITICAL: median booking lead time is only 9-11 days in Jul/Aug — over 40% of August revenue books within 7 days of arrival.
  Thin forward pace for July/August when looking from May or June is structurally NORMAL, not a demand problem. Never recommend discounting peak months based on thin pace more than 2 weeks out.
- April books furthest out (37-day median) — the one month where early-bird pricing has logic
- Weekend ADR runs $60-90 above weekday in peak months, widening to $84 in October (foliage)
- November/December are viable winter weekends — NYC/Philly short-getaway feeder market
- Do not recommend discounting weekends

Write one action sentence for the marketing team. Be specific about which week or weekend needs attention
and what action to take (e.g. push organic content, email past guests, boost paid spend). If the most
urgent pace gap is in a last-minute-booking month (Jun/Jul/Aug) and weeks-out > 2, do not treat it as urgent —
say so explicitly and point to the next genuinely actionable item instead."""

        try:
            payload = json.dumps({
                "model": "claude-sonnet-4-6",
                "max_tokens": 200,
                "messages": [{"role": "user", "content": prompt}]
            }).encode()
            req = urllib.request.Request(
                "https://api.anthropic.com/v1/messages",
                data=payload,
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01"
                }
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                result = json.loads(resp.read())
                text = result["content"][0]["text"].strip()
                print("  ✅ AI action item generated")
                return text
        except Exception as e:
            print(f"  ⚠️  Claude API call failed: {e} — using heuristic action item")

    # Heuristic fallback — suppress false urgency for last-minute-booking months
    if worst["flag"] == "🔴" and not (worst["profile"] == "last_minute" and worst["weeks_out"] > 2):
        return (f"{worst['week_label']} is the most urgent gap at {worst['booked_pct']:.0%} booked vs "
                f"{worst['target_pct']:.0%} target — push organic content and email past guests this week.")
    urgent_wknd = next((w for w in weekend_rows if w["flag"] == "🔴"), None)
    if urgent_wknd:
        return (f"{urgent_wknd['label']} is under-booked — boost paid spend targeting "
                f"NYC/Philly last-minute drive-market guests.")
    return "No urgent pace or weekend gaps this week — hold steady and monitor."


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5: Post Slack digest
# ─────────────────────────────────────────────────────────────────────────────
def post_slack_digest(ytd_data, pace_data, as_of_date):
    import urllib.request

    webhook_url = os.environ.get("SLACK_WEBHOOK_URL_PROMISED_LAND", "")
    dry_run = os.environ.get("DRY_RUN", "").lower() in ("1", "true", "yes")
    today = date.today()
    fiscal_year = CONFIG["fiscal_year"]
    # The YTD file ends the day before today and the pace file starts today —
    # any week that straddles that boundary (e.g. this week, Mon-Sun) needs
    # both sources merged, or the days that already closed this week silently
    # drop out of the pace/pickup/YoY tables. Only pull in the *recent* tail
    # of YTD (last 10 days) — merging the full YTD history would make every
    # day since January look like "new pickup" against a snapshot that never
    # tracked those already-closed months.
    ytd_daily_all = _flatten_daily_rows(ytd_data)
    recent_ytd_daily = {}
    for d, row in ytd_daily_all.items():
        try:
            m, dd = map(int, d.split("/"))
            row_date = date(fiscal_year, m, dd)
        except (ValueError, AttributeError):
            continue
        if 0 <= (today - row_date).days <= 10:
            recent_ytd_daily[d] = row
    pace_daily = {**recent_ytd_daily, **_flatten_daily_rows(pace_data)}

    prev_snapshot = _load_pace_snapshot(CONFIG["snapshot_cache"])
    pickup_stats = _compute_pickup_stats(pace_daily, prev_snapshot, today, fiscal_year)
    if dry_run:
        print("  ℹ️  DRY_RUN set — not overwriting pickup snapshot cache")
    else:
        _save_pace_snapshot(CONFIG["snapshot_cache"], pace_daily, as_of_date)

    if pickup_stats:
        pickup_text = (
            f"• *{pickup_stats['pickup_7d']:+d} nights* net pickup in last 7 days"
            + (f" · ADR on new bookings: *${pickup_stats['adr_new']:.0f}*"
               if pickup_stats["adr_new"] > 0 else "")
            + f"\n• *{pickup_stats['pickup_14d']} of those* for arrivals within the next 14 days"
            + f"\n• *${pickup_stats['revenue_pickup']:+,.0f}* net revenue picked up vs last week"
        )
    else:
        pickup_text = "• First snapshot recorded — pickup comparison starts next week"

    py_daily = parse_prior_year_daily(CONFIG["py_file"])
    yoy_rows = _build_yoy_week_comparison(ytd_daily_all, py_daily, today) if py_daily else []
    if yoy_rows:
        yoy_lines = []
        for r in yoy_rows:
            adr_str = f"{r['adr_yoy_pct']:+.0%} ADR" if r["adr_yoy_pct"] is not None else "ADR n/a"
            line = (f"{r['week_label']:<10} —  {r['cur_nights']:>2} nights "
                    f"vs {r['py_nights']:>2} LY ({r['nights_yoy']:+d})  ·  {adr_str}")
            if r["anomaly"]:
                line += f"  ⚠️ LY anomaly: {r['anomaly']}"
            yoy_lines.append(line)
        yoy_text = "\n".join(yoy_lines)
    else:
        yoy_text = None

    pace_rows = _build_forward_pace_table(pace_daily, today)
    push_now_idx = next(
        (i for i, r in enumerate(pace_rows[:2])
         if r["flag"] == "🔴" and not (r["profile"] == "last_minute" and r["weeks_out"] > 2)),
        None
    )

    pace_lines = []
    for idx, r in enumerate(pace_rows):
        if r["weeks_out"] == 5:
            line = (f"{r['flag']}  {r['week_label']:<10} —  "
                    f"{r['booked']:>2}/{r['avail']} ({r['booked_pct']:.0%})  —")
        else:
            line = (f"{r['flag']}  {r['week_label']:<10} —  "
                    f"{r['booked']:>2}/{r['avail']} ({r['booked_pct']:.0%})  "
                    f"tgt {r['target_pct']:.0%}")
            if idx == push_now_idx:
                line += "  ← push now"
        pace_lines.append(line)
    pace_text = "\n".join(pace_lines)

    weekend_rows = _build_weekend_watch(pace_daily, today)
    weekend_lines = [
        f"{w['flag']} {w['label']} — {w['weeks_out']}wk{'s' if w['weeks_out'] != 1 else ''} out · "
        f"Fri {w['fri_booked']}/{w['fri_units']} · Sat {w['sat_booked']}/{w['sat_units']} · {w['avg_pct']:.0%} booked"
        for w in weekend_rows
    ]
    weekend_text = "\n".join(weekend_lines)

    action_text = _get_action_item(pickup_stats, pace_rows, weekend_rows)

    blocks = [
        {
            "type": "header",
            "text": {"type": "plain_text", "text": f"🏔 Promised Land — Monday Digest · {as_of_date}"}
        },
        {"type": "divider"},
        {
            "type": "section",
            "text": {"type": "mrkdwn", "text": f"📊 *This week's pickup*\n{pickup_text}"}
        },
        {"type": "divider"},
        {
            "type": "section",
            "text": {"type": "mrkdwn", "text": f"📅 *Forward pace — next 5 weeks*\n```{pace_text}```"}
        },
        {"type": "divider"},
    ] + ([{
            "type": "section",
            "text": {"type": "mrkdwn", "text": f"📈 *Closed weeks vs last year (actuals)*\n```{yoy_text}```"}
        },
        {"type": "divider"}] if yoy_text else []) + [
        {
            "type": "section",
            "text": {"type": "mrkdwn", "text": f"🏖 *Weekend watch*\n{weekend_text}"}
        },
        {"type": "divider"},
        {
            "type": "section",
            "text": {"type": "mrkdwn", "text": f"🎯 *This week:* {action_text}"}
        },
        {
            "type": "context",
            "elements": [
                {
                    "type": "mrkdwn",
                    "text": (f"<{CONFIG['full_report_url']}|View Full Report>  ·  "
                             f"Promised Land Pipeline · {as_of_date}")
                }
            ]
        }
    ]

    if not webhook_url:
        print("\n📣 Slack webhook not configured — printing digest preview instead")
        print("   Set SLACK_WEBHOOK_URL_PROMISED_LAND environment variable to enable posting")
        print(f"\n{'='*60}")
        print(f"Promised Land — Monday Digest · {as_of_date}")
        print(f"{'='*60}")
        print(f"\nThis week's pickup:\n{pickup_text}")
        print(f"\nForward pace — next 5 weeks:\n{pace_text}")
        if yoy_text:
            print(f"\nClosed weeks vs last year (actuals):\n{yoy_text}")
        print(f"\nWeekend watch:\n{weekend_text}")
        print(f"\nThis week: {action_text}")
        return

    try:
        payload = json.dumps({"blocks": blocks}).encode()
        req = urllib.request.Request(
            webhook_url,
            data=payload,
            headers={"Content-Type": "application/json"}
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            print(f"  ✅ Slack digest posted (status {resp.status})")
    except Exception as e:
        print(f"  ⚠️  Slack post failed: {e}")


def main():
    print("=" * 60)
    print("  PROMISED LAND DATA PIPELINE")
    print(f"  Property: {CONFIG['property_code']}")
    print(f"  Run date: {date.today().strftime('%Y-%m-%d')}")
    print("=" * 60)

    ytd_data = parse_cloudbeds_export(CONFIG["ytd_file"])
    pace_data = parse_cloudbeds_export(CONFIG["pace_file"])

    post_slack_digest(ytd_data, pace_data, as_of_date=date.today().strftime("%Y-%m-%d"))

    print("\n✅ Pipeline complete.")
    print("=" * 60)

    return {"ytd": ytd_data, "pace": pace_data}


if __name__ == "__main__":
    main()
